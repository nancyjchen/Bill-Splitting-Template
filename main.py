from openpyxl import Workbook

def main(): 
    workbook = Workbook()
    filename = "output.xlsx"
    
    create_sheets(workbook)
    num_people, ppl_names = retrieve_infO()
    
    update_overall(workbook)
    start_idx, final_idx = update_items(workbook, num_people, ppl_names)
    update_readme(workbook, start_idx, final_idx, ppl_names)

    ### save workbook 
    workbook.save(filename)

### create sheets
def create_sheets(workbook): 
    workbook.active.title = "README"
    workbook.create_sheet('Items')
    workbook.create_sheet('Overall')

### retrieve number of people and their names 
def retrieve_infO(): 
    num_people = int(input("Enter number of people splitting the bill: "))
    ppl_names = []
    for i in range(1, num_people + 1):
        ppl_names.append(input("Enter name of Person " + str(i) + ": "))
    return num_people, ppl_names
    
### update 'Overall' sheet
# set data 
def update_overall(workbook):
    overall_sheet = workbook['Overall']
    overall_items = ["Subtotal", "Tax", "Total", "Tips", "Total Paid", "", "% Increase"]
    count = 1
    for item in overall_items: 
        overall_sheet["A" + str(count)] = item
        count += 1
    
    overall_sheet["B7"] = "= B5/B1"
    

### update 'Items' sheet 
def update_items(workbook, num_people, ppl_names):
    item_sheet = workbook["Items"]
    item_sheet["A1"] = "Item"
    item_sheet["B1"] = "Cost"
    item_sheet["C1"] = "# Items"
    count = 1
    start_idx = ord('D')
    for i in range(start_idx, start_idx + num_people):
        item_sheet[chr(i) + "1"] = "(" + str(count) + ") " + ppl_names[count - 1]
        count += 1

    cost_idx = start_idx + num_people + 1
    item_sheet[chr(cost_idx) + "1"] = "Total Cost" 
    ppl_idx = cost_idx + 1
    item_sheet[chr(ppl_idx) + "1"] = "# People" 
    for i in range(2, 100): 
        item_sheet[chr(cost_idx) + str(i)] = '= B' + str(i) + '*C' + str(i)
        item_sheet[chr(ppl_idx) + str(i)] = '=COUNTIF(D' + str(i) + ':' + \
            chr(ppl_idx - 2) + str(i) + ', TRUE)'

    split_idx = ppl_idx + 2
    count = 1
    for i in range(split_idx, split_idx + num_people):
        item_sheet[chr(i) + "1"] = "(" + str(count) + ") " + ppl_names[count - 1]
        for j in range(2, 100): 
            item_sheet[chr(i) + str(j)] = \
                '=IF(' + chr(start_idx + count - 1) + str(j) + '=TRUE, ' + \
                chr(cost_idx) + str(j) + '/' + chr(ppl_idx) + str(j) + ', 0' + ')'
        count += 1

    final_idx = split_idx + num_people + 1
    count = 1
    for i in range(final_idx, final_idx + num_people):
        item_sheet[chr(i) + "1"] = "(" + str(count) + ") " + ppl_names[count - 1]
        item_sheet[chr(i) + "2"] = \
                '=SUM(' + chr(split_idx + count - 1) + '2:' + \
                chr(split_idx + count - 1) + '100)'
        count += 1
    
    return start_idx, final_idx


### update 'README' sheet 
def update_readme(workbook, start_idx, final_idx, ppl_names):
    rm_sheet = workbook["README"]
    rm_sheet["A1"] = "Name"
    rm_sheet["B1"] = "Total Amt"

    # update table 
    count = 2
    for name in ppl_names:
        rm_sheet["A" + str(count)] = name
        rm_sheet["B" + str(count)] = "=Overall!B7 * SUM(Items!" + \
            chr(final_idx + count - 2) + "2:Items!" + \
            chr(final_idx + count - 2) + "100)"
        count += 1
    
    # update notes
    rm_sheet["D1"] = "Update columns A" + " to " + chr(start_idx + len(ppl_names) - 1) + \
        " in 'Items' and rows 2 to 6 in 'Overall'"
    
if __name__ == "__main__":
    main()